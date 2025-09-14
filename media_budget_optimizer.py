#!/usr/bin/env python3
"""
Media Budget Optimization Excel Generator
Creates an Excel file with pre-configured optimization model for media budget allocation.
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.datavalidation import DataValidation
import os

def create_media_budget_excel():
    """Create Excel file with media budget optimization model"""
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Media Budget Optimization"
    
    # Define the data structure
    data = [
        ["Decision Variables", ""],
        ["Social media $ (S)", 1000],
        ["Television $ (T)", 1000],
        ["", ""],
        ["Total Spend", "=B2 + B3"],
        ["Total Reach (people)", "=50*B2 + 40*B3"],
        ["", ""],
        ["Budget available ($)", 50000],
        ["Max TV percent", 0.7],
        ["Max TV $ limit", "=B8 * B9"],
        ["Required Reach (people)", 1500000],
        ["", ""],
        ["Model A Results (Maximize Reach)", ""],
        ["Optimal Social $", "=B2"],
        ["Optimal TV $", "=B3"],
        ["Max Reach", "=B6"],
        ["", ""],
        ["Model B Results (Minimize Spend)", ""],
        ["Min Social $", "=B2"],
        ["Min TV $", "=B3"],
        ["Min Total Spend", "=B5"],
        ["Achieved Reach", "=B6"]
    ]
    
    # Add data to worksheet
    for row_idx, (label, value) in enumerate(data, 1):
        ws[f'A{row_idx}'] = label
        ws[f'B{row_idx}'] = value
    
    # Format the worksheet
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15
    
    # Add headers and formatting
    header_rows = [1, 7, 12, 17]  # Rows with section headers
    for row in header_rows:
        ws[f'A{row}'].font = ws[f'A{row}'].font.copy(bold=True)
        ws[f'A{row}'].fill = ws[f'A{row}'].fill.copy(fgColor="CCCCCC")
    
    # Add borders to important cells
    from openpyxl.styles import Border, Side
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Apply borders to data cells
    for row in range(2, 6):  # Decision variables and totals
        for col in ['A', 'B']:
            ws[f'{col}{row}'].border = thin_border
    
    for row in range(8, 11):  # Parameters
        for col in ['A', 'B']:
            ws[f'{col}{row}'].border = thin_border
    
    # Create a second sheet with solver instructions
    ws2 = wb.create_sheet("Solver Instructions")
    
    instructions = [
        ["SOLVER SETUP INSTRUCTIONS", ""],
        ["", ""],
        ["MODEL A - MAXIMIZE REACH", ""],
        ["1. Go to Data > Solver", ""],
        ["2. Set Objective: B6", ""],
        ["3. To: Max", ""],
        ["4. By Changing: B2:B3", ""],
        ["5. Add Constraints:", ""],
        ["   - B2 + B3 = B8 (use all budget)", ""],
        ["   - B3 <= B10 (TV limit)", ""],
        ["   - B2 >= 0", ""],
        ["   - B3 >= 0", ""],
        ["6. Solving Method: Simplex LP", ""],
        ["7. Click Solve", ""],
        ["", ""],
        ["MODEL B - MINIMIZE SPEND", ""],
        ["1. Go to Data > Solver", ""],
        ["2. Set Objective: B5", ""],
        ["3. To: Min", ""],
        ["4. By Changing: B2:B3", ""],
        ["5. Add Constraints:", ""],
        ["   - 50*B2 + 40*B3 >= B11 (reach target)", ""],
        ["   - B3 <= B10 (TV limit)", ""],
        ["   - B2 >= 0", ""],
        ["   - B3 >= 0", ""],
        ["6. Solving Method: Simplex LP", ""],
        ["7. Click Solve", ""],
        ["", ""],
        ["EXPECTED RESULTS", ""],
        ["Model A: Social=$50,000, TV=$0, Reach=2,500,000", ""],
        ["Model B: Social=$30,000, TV=$0, Spend=$30,000", ""]
    ]
    
    for row_idx, (instruction, value) in enumerate(instructions, 1):
        ws2[f'A{row_idx}'] = instruction
        ws2[f'B{row_idx}'] = value
    
    ws2.column_dimensions['A'].width = 50
    ws2.column_dimensions['B'].width = 30
    
    # Format instruction headers
    header_rows = [1, 3, 17, 25]
    for row in header_rows:
        ws2[f'A{row}'].font = ws2[f'A{row}'].font.copy(bold=True)
        ws2[f'A{row}'].fill = ws2[f'A{row}'].fill.copy(fgColor="FFFF00")
    
    return wb

def create_csv_import_file():
    """Create CSV file that can be imported into Excel"""
    
    data = {
        'Label': [
            'Decision Variables',
            'Social media $ (S)',
            'Television $ (T)',
            '',
            'Total Spend',
            'Total Reach (people)',
            '',
            'Budget available ($)',
            'Max TV percent',
            'Max TV $ limit',
            'Required Reach (people)',
            '',
            'Model A Results (Maximize Reach)',
            'Optimal Social $',
            'Optimal TV $',
            'Max Reach',
            '',
            'Model B Results (Minimize Spend)',
            'Min Social $',
            'Min TV $',
            'Min Total Spend',
            'Achieved Reach'
        ],
        'Value_Formula': [
            '',
            '1000',
            '1000',
            '',
            '=B2 + B3',
            '=50*B2 + 40*B3',
            '',
            '50000',
            '0.7',
            '=B8 * B9',
            '1500000',
            '',
            '',
            '=B2',
            '=B3',
            '=B6',
            '',
            '',
            '=B2',
            '=B3',
            '=B5',
            '=B6'
        ]
    }
    
    df = pd.DataFrame(data)
    return df

def main():
    """Main function to create both Excel and CSV files"""
    
    print("Creating Media Budget Optimization files...")
    
    # Create Excel file
    try:
        wb = create_media_budget_excel()
        excel_path = "/Users/jasminpingol/Downloads/thesis75/nutrisaur11/media_budget_optimization.xlsx"
        wb.save(excel_path)
        print(f"✅ Excel file created: {excel_path}")
    except Exception as e:
        print(f"❌ Error creating Excel file: {e}")
    
    # Create CSV file
    try:
        df = create_csv_import_file()
        csv_path = "/Users/jasminpingol/Downloads/thesis75/nutrisaur11/media_budget_optimization.csv"
        df.to_csv(csv_path, index=False)
        print(f"✅ CSV file created: {csv_path}")
    except Exception as e:
        print(f"❌ Error creating CSV file: {e}")
    
    print("\n" + "="*50)
    print("FILES CREATED SUCCESSFULLY!")
    print("="*50)
    print("1. Excel file: media_budget_optimization.xlsx")
    print("   - Ready to use with pre-configured formulas")
    print("   - Includes solver instructions sheet")
    print("   - Just open and run Solver!")
    print()
    print("2. CSV file: media_budget_optimization.csv")
    print("   - Import into Excel if needed")
    print("   - Copy column A to Excel column A")
    print("   - Copy column B to Excel column B")
    print()
    print("QUICK START:")
    print("1. Open the Excel file")
    print("2. Go to Data > Solver")
    print("3. Follow instructions in the 'Solver Instructions' sheet")
    print("4. Run optimization!")

if __name__ == "__main__":
    main()
